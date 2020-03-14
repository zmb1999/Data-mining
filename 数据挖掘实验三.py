import numpy as np
import math
from sklearn import preprocessing
from sklearn.model_selection import RepeatedKFold,StratifiedKFold
import pandas
from scipy.io import arff
from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import LinearSVC
from sklearn import metrics
from sklearn.model_selection import GridSearchCV, train_test_split
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")

file_home = r'C:\Users\99769\PycharmProjects\untitled\result.xlsx'

wb = load_workbook(filename=file_home)
sheet_ranges = wb['Sheet1']
ws = wb['Sheet1']

def getdata(data):
    instance = data.shape[0]
    attribute = data.shape[1] - 1
    Defective = data[:, attribute]
    data = np.delete(data, attribute, axis=1)
    return data,instance,attribute,Defective

def getfeature(feature,fRank,n):
    Data = []
    for i in range(n):
         Data.append(feature[fRank[i]])
    return Data

def getDefectiveNum(Defective):
    DefectiveNum = []
    for i,v in enumerate(Defective):
        if v == b'Y':
            DefectiveNum.append(1)
        else:
            DefectiveNum.append(0)
    DefectiveNum = [float(i) for i in DefectiveNum]
    return DefectiveNum

def getminposition(mylist):
    minnum = min(mylist)
    position = mylist.index(minnum)
    return position

def delete(list1,list2):
    deltalist = []
    for i,v in enumerate(list1):
        deltalist.append(abs(list1[i]-list2[i]))
    return deltalist

def add(list1,list2):
    sumlist = []
    for i,v in enumerate(list1):
        sumlist.append(abs(list1[i]+list2[i]))
    return sumlist

def DecisionTree(X_train,Y_train):
    tree = DecisionTreeClassifier(max_depth=5, random_state=0)
    tree.fit(X_train, Y_train)
    return tree

def SVM(X_train, Y_train):
    lsvc = LinearSVC(max_iter=10000)
    lsvc.fit(X_train, Y_train)
    return lsvc

def score(y_true,y_pred):
    # TP = np.sum(np.multiply(y_true, y_pred))
    # # false positive
    # FP = np.sum(np.logical_and(np.equal(y_true, 0), np.equal(y_pred, 1)))
    # # false negative
    # FN = np.sum(np.logical_and(np.equal(y_true, 1), np.equal(y_pred, 0)))
    # # true negative
    # TN = np.sum(np.logical_and(np.equal(y_true, 0), np.equal(y_pred, 0)))
    # P = TP/(TP+FP)
    # R = TP/(TP+FN)
    # F = 2 * P * R / (P + R)
    auc_score = metrics.roc_auc_score(y_true, y_pred)
    F = metrics.f1_score(y_true, y_pred, average='binary', pos_label=1)
    return F, auc_score

import os
DataPath = './Data/'
fileList = os.listdir(DataPath)

cnt = 4
for filename in fileList:
    dataset, mate = arff.loadarff(DataPath + filename)
    #print(dataset)
    df = pandas.DataFrame(dataset)
    #print(df)
    #dataset_ls = list(dataset)
    originData = np.array(df)
    #print(originData.shape)
    data, instance, attribute, Defective = getdata(originData)
    # print(data, "\n",instance,"\n", attribute,"\n", Defective)
    #print(Tdata)
    DefectiveNum = getDefectiveNum(Defective)
    # print(DefectiveNum)
    #标准化
    data = data.astype('float64')
    scaled = preprocessing.scale(data)
    # print(scaled)

    #计算k
    k = np.sum(Defective == b'N')/np.sum(Defective == b'Y')
    k = int(k)

    # print(k)
    #标准化距离
    ed = np.zeros((instance, instance),dtype='float32')
    alldata = instance*instance
    for i in range(0, instance):
        for j in range(0, instance):
            ed[i][j] = np.linalg.norm(scaled[i,:] - scaled[j,:])
    # print(ed)

    #临近样本
    rank = []
    for i in range(0, instance):
        rank.append(pandas.Series(ed[i, :]).rank(method='min').tolist())
    #print(rank)

    nearest = []

    for index,i in enumerate(rank):
        n = []
        num = 0
        while 1:
            position = getminposition(i)
            if Defective[position] == b'Y' or position == index:
                i[position] = max(i)
            else:
                n.append(position)
                i[position] = max(i)
                num += 1
            if num == k:
                break
        nearest.append(n)
    # print(nearest)

    #特征差值
    delta = []
    for i,v in enumerate(data):
            d = []
            for j,w in enumerate(nearest[i]):
                d.append(delete(data[i], data[w]))
            delta.append(d)
    #print(delta)

    #特征权重
    W = np.zeros(attribute)
    #print(W)
    for i,v in enumerate(delta):
        if Defective[i] == b'Y':
            for j in v:
                W = add(W,pandas.Series(j).rank(method='min').tolist())
    # print(W)

    #特征排序列表
    WRank = pandas.Series(W).rank(method='min').tolist()
    fRank = []
    flag = 0
    while 1:
        for i, v in enumerate(WRank):
            if v == max(WRank):
                fRank.append(i)
                WRank[i] = -1
                flag += 1
            if flag == attribute:
                break
        if flag == attribute:
            break

    print(filename+"特征排序列表:",fRank)

    skf = StratifiedKFold(n_splits=10, random_state=None, shuffle=True)
    # skf = RepeatedKFold(n_splits=10, n_repeats=10, random_state=0)
    # skf.get_n_splits(data, DefectiveNum)
    # print(skf)
    DefectiveNum = np.array(DefectiveNum)
    sum1, sum2, sum3, sum4, sum5, sum6, sum7, sum8 = 0, 0, 0, 0, 0, 0, 0, 0
    for train_index, test_index in skf.split(scaled, DefectiveNum):
        print('train_index', train_index, 'test_index', test_index)
        # print(type(train_index[1]))
        # print("TRAIN:", train_index,'\n', "TEST:", test_index)
        All_X_train, All_X_test = scaled[train_index], scaled[test_index]
        Y_train, Y_test = np.array(DefectiveNum)[train_index], DefectiveNum[test_index]
        log2_X_train,log2_X_test = [], []
        for i in range(len(All_X_train)):
            log2_X_train.append(getfeature(All_X_train[i],fRank,int(math.log2(All_X_train.shape[1]))))
        for i in range(len(All_X_test)):
            log2_X_test.append(getfeature(All_X_test[i],fRank,int(math.log2(All_X_test.shape[1]))))
        # print(len(log2_X_train),len(log2_X_train[0]),len(log2_X_test),len(log2_X_test[0]))
        log2_tree = DecisionTree(log2_X_train,Y_train)
        All_tree = DecisionTree(All_X_train,Y_train)

        log2_SVM = SVM(log2_X_train,Y_train)
        All_SVM = SVM(All_X_train,Y_train)

        log2_tree_pred = log2_tree.predict(log2_X_test)
        All_tree_pree = All_tree.predict(All_X_test)

        log2_SVM_pred = log2_SVM.predict(log2_X_test)
        All_SVM_pred = All_SVM.predict(All_X_test)

        F_log2_tree, auc_score_log2_tree = score(Y_test,log2_tree_pred)
        F_All_tree, auc_score_All_tree = score(Y_test,All_tree_pree)

        F_log2_SVM, auc_score_log2_SVM = score(Y_test,log2_SVM_pred)
        F_All_SVM, auc_score_All_SVM = score(Y_test,All_tree_pree)

        sum1 += F_log2_tree
        sum2 += F_All_tree
        sum3 += F_log2_SVM
        sum4 += F_All_SVM
        sum5 += auc_score_log2_tree
        sum6 += auc_score_All_tree
        sum7 += auc_score_log2_SVM
        sum8 += auc_score_All_SVM

        print(
              filename+"决策树log2子集F1值:",F_log2_tree,"决策树原始特征集F1值:",F_All_tree,'\n',
              filename+"SVM训练log2子集F1值:",F_log2_SVM,"SVM训练原始特征集F1值:",F_All_SVM,'\n',
              filename + "决策树log2子集AUC:", auc_score_log2_tree, "决策树原始特征集AUC:", auc_score_All_tree, '\n',
              filename + "SVM训练log2子集AUC:", auc_score_log2_SVM, "SVM训练原始特征集AUC:", auc_score_All_SVM
              )
    ws['A' + str(cnt)] = filename
    ws['B' + str(cnt)] = sum1 / 10
    ws['C' + str(cnt)] = sum2 / 10
    ws['F' + str(cnt)] = sum3 / 10
    ws['G' + str(cnt)] = sum4 / 10
    ws['D' + str(cnt)] = sum5 / 10
    ws['E' + str(cnt)] = sum6 / 10
    ws['H' + str(cnt)] = sum7 / 10
    ws['I' + str(cnt)] = sum8 / 10
    wb.save(file_home)
    cnt += 1

